# -*- coding: utf-8 -*-
"""
MASTER DATA OLUŞTURMA ARACI - TAM OTOMATİK
Klasördeki tüm Excel dosyalarını tarar, hafta/temsilci/tip otomatik algılar,
veriyi çeker ve MASTER_DATA.xlsx oluşturur.

Kullanım:
  python master_data_olustur.py          # Normal çalıştır
  python master_data_olustur.py --izle   # Dosya değişikliği izle (watch mode)
"""

import re
import sys
import io
import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

# Windows encoding fix
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE_DIR = Path(__file__).parent

# ============================================================
# LOGLAMA
# ============================================================
LOG_FILE = BASE_DIR / 'master_data.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger('master_data')

# ============================================================
# AYARLAR - Gerekirse buradan düzenleyin
# ============================================================
DOSYA_DIZINI = BASE_DIR          # Excel dosyalarının bulunduğu klasör
CIKTI_DOSYA = 'MASTER_DATA.xlsx' # Çıktı dosya adı
YEDEK_FILTRELE = True            # _YEDEK dosyaları atla
# ============================================================

# Stiller
H_FILL = PatternFill('solid', fgColor='1F4E79')
H_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
D_FONT = Font(name='Arial', size=10)
B_FONT = Font(name='Arial', size=10, bold=True)
B = Border(*(Side(style='thin', color='B0B0B0'),) * 4)
A_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_LEFT = Alignment(vertical='center', wrap_text=True)
WARN_FILL = PatternFill('solid', fgColor='FFF3CD')
ERR_FILL = PatternFill('solid', fgColor='F8D7DA')
OK_FILL = PatternFill('solid', fgColor='D4EDDA')
BLUE_FILL = PatternFill('solid', fgColor='D6EAF8')
LIGHT_FILL = PatternFill('solid', fgColor='FFF5F5')
INFO_FILL = PatternFill('solid', fgColor='E3F2FD')
PURPLE_FILL = PatternFill('solid', fgColor='EDE7F6')
AUTO_FILL = PatternFill('solid', fgColor='E8F5E9')

TR_MONTHS = {
    'OCAK': 1, 'SUBAT': 2, 'MART': 3, 'NISAN': 4,
    'MAYIS': 5, 'HAZIRAN': 6, 'TEMMUZ': 7, 'AGUSTOS': 8,
    'EYLUL': 9, 'EKIM': 10, 'KASIM': 11, 'ARALIK': 12,
}
# Doğru Türkçe ay isimleri (görüntüleme için)
TR_MONTH_DISPLAY = {
    1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan',
    5: 'Mayıs', 6: 'Haziran', 7: 'Temmuz', 8: 'Ağustos',
    9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık',
}
TR_DAYS = {0: 'Pazartesi', 1: 'Salı', 2: 'Çarşamba', 3: 'Perşembe', 4: 'Cuma', 5: 'Cumartesi', 6: 'Pazar'}


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = A_CENTER
        cell.border = B


def style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = D_FONT
        cell.border = B
        cell.alignment = A_LEFT
        if fill:
            cell.fill = fill


def normalize_tr(s):
    return s.upper().replace('İ', 'I').replace('Ş', 'S').replace('Ğ', 'G').replace('Ü', 'U').replace('Ö', 'O').replace('Ç', 'C')


def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


def safe_date(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d.%m.%Y')
    s = str(v).strip()
    if not s:
        return ''
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f'{m.group(3)}.{m.group(2)}.{m.group(1)}'
    return s


# ============================================================
# OTOMATİK ALGILAMA FONKSİYONLARI
# ============================================================

def dosya_listesi_al(dizin):
    """xlsx dosyalarını filtrele ve listele."""
    files = []
    for f in sorted(dizin.glob('*.xlsx')):
        fn = f.name
        if YEDEK_FILTRELE and '_YEDEK' in fn:
            continue
        if fn.upper().startswith('MASTER'):
            continue
        if fn.startswith('~$'):
            continue
        files.append(fn)
    return files


def hafta_algila(dosya_listesi):
    """Dosya isimlerinden hafta tarih aralıklarını otomatik algıla."""
    haftalar = {}  # {(start_date, end_date): label}

    # Önce tam tarihli dosyalardan yılı tespit et (Pattern 2 fallback için)
    detected_years = set()
    for fn in dosya_listesi:
        m = re.search(r'\d{2}\.\d{2}\.(\d{4})', fn)
        if m:
            detected_years.add(int(m.group(1)))
    default_year = max(detected_years) if detected_years else datetime.now().year

    for fn in dosya_listesi:
        fn_norm = normalize_tr(fn)

        # Pattern 1: DD.MM.YYYY-DD.MM.YYYY
        m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})-(\d{2})\.(\d{2})\.(\d{4})', fn)
        if m:
            sd = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            ed = datetime(int(m.group(6)), int(m.group(5)), int(m.group(4)))
            key = (sd, ed)
            if key not in haftalar:
                month_name = TR_MONTH_DISPLAY.get(sd.month, str(sd.month))
                label = f'{sd.day:02d}-{ed.day:02d} {month_name} {sd.year}'
                haftalar[key] = label
            continue

        # Pattern 2: DD-DD AY (e.g. "26-30 OCAK")
        m2 = re.search(r'(\d{1,2})-(\d{1,2})\s+([A-Z]+)', fn_norm)
        if m2:
            day_s, day_e, month_str = int(m2.group(1)), int(m2.group(2)), m2.group(3)
            month_num = TR_MONTHS.get(month_str)
            if month_num:
                # Yılı tahmin et: dosya adında yıl varsa onu al, yoksa diğer dosyalardan algılanan yılı kullan
                y_m = re.search(r'20\d{2}', fn)
                year = int(y_m.group()) if y_m else default_year
                sd = datetime(year, month_num, day_s)
                ed = datetime(year, month_num, day_e)
                key = (sd, ed)
                if key not in haftalar:
                    month_name = TR_MONTH_DISPLAY.get(month_num, month_str.capitalize())
                    label = f'{day_s:02d}-{day_e:02d} {month_name} {year}'
                    haftalar[key] = label

    # Sırala ve döndür: [(label, start_str, end_str, start_dt, end_dt), ...]
    result = []
    for (sd, ed) in sorted(haftalar.keys()):
        label = haftalar[(sd, ed)]
        result.append((label, sd.strftime('%d.%m.%Y'), ed.strftime('%d.%m.%Y'), sd, ed))
    return result


def temsilci_algila(dosya_listesi):
    """Dosya isimlerinden temsilci adlarını otomatik algıla."""
    # Bilinen dosya tipi kelimeleri (bunları temsilci adından çıkar)
    tip_kelimeleri = {
        'PLANLANAN', 'YAPILAN', 'ZIYARET', 'SIPARIS', 'FORMU', 'HAFTALIK',
        'PLANI', 'OCAK', 'SUBAT', 'MART', 'NISAN', 'MAYIS', 'HAZIRAN',
        'TEMMUZ', 'AGUSTOS', 'EYLUL', 'EKIM', 'KASIM', 'ARALIK',
        'XLSX', 'YEDEK',
    }

    temsilciler = set()

    for fn in dosya_listesi:
        # Dosya adından tarih ve uzantıyı temizle
        name = fn.replace('.xlsx', '').replace('.XLSX', '')
        name = re.sub(r'\d{2}\.\d{2}\.\d{4}-\d{2}\.\d{2}\.\d{4}', '', name)
        name = re.sub(r'\d{1,2}-\d{1,2}', '', name)
        name = name.replace('_YEDEK', '').replace('_yedek', '')

        # Kalan kelimeleri parçala
        parts = name.strip().split()
        # Tip kelimelerini çıkar
        name_parts = []
        for p in parts:
            p_norm = normalize_tr(p)
            if p_norm not in tip_kelimeleri and len(p) > 1:
                name_parts.append(p)

        # En az 2 kelimelik bir isim olmalı (Ad Soyad)
        if len(name_parts) >= 2:
            # Son 2-3 kelime genelde isim
            candidate = ' '.join(name_parts).strip()
            if candidate:
                temsilciler.add(candidate)

    # Normalize edip birleştir (aynı kişinin farklı yazımlarını)
    merged = {}
    for t in temsilciler:
        key = normalize_tr(t).replace(' ', '')
        if key not in merged or len(t) > len(merged[key]):
            merged[key] = t

    return sorted(merged.values())


def dosya_tipi_bul(fn):
    """Dosya tipini otomatik algıla."""
    fn_norm = normalize_tr(fn)
    if 'PLANLANAN' in fn_norm or 'ZIYARET PLANI' in fn_norm:
        return 'Planlanan Ziyaret'
    elif 'YAPILAN' in fn_norm:
        return 'Yapılan Ziyaret'
    elif 'SIPARIS' in fn_norm:
        return 'Sipariş Formu'
    return None


def temsilci_bul(fn, temsilciler):
    """Dosya adından temsilciyi bul."""
    fn_norm = normalize_tr(fn).replace(' ', '').replace('_', '')
    for rep in temsilciler:
        if normalize_tr(rep).replace(' ', '') in fn_norm:
            return rep
    return ''


def hafta_bul(fn, haftalar):
    """Dosya adından haftayı bul."""
    fn_norm = normalize_tr(fn)
    # Önce tam tarih eşleştir
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})-(\d{2})\.(\d{2})\.(\d{4})', fn)
    if m:
        sd_str = f'{m.group(1)}.{m.group(2)}.{m.group(3)}'
        ed_str = f'{m.group(4)}.{m.group(5)}.{m.group(6)}'
        for label, ws, we, _, _ in haftalar:
            if sd_str == ws and ed_str == we:
                return label

    # Sonra kısa pattern (DD-DD AY)
    for label, ws, we, sd, ed in haftalar:
        s_d, e_d = sd.day, ed.day
        month_name_norm = normalize_tr(TR_MONTH_DISPLAY.get(sd.month, ''))
        # Hem padded (02-06) hem unpadded (2-6) dene
        patterns = [f'{s_d}-{e_d}', f'{s_d:02d}-{e_d:02d}']
        for short_pat in patterns:
            if short_pat in fn and month_name_norm in fn_norm:
                return label

    return ''


# ============================================================
# ANA VERİ ÇEKME
# ============================================================

def master_data_olustur():
    xlsx_files = dosya_listesi_al(DOSYA_DIZINI)
    if not xlsx_files:
        log.error("Hiç Excel dosyası bulunamadı!")
        log.error(f"  Dizin: {DOSYA_DIZINI}")
        return False

    log.info(f"Taranan dizin: {DOSYA_DIZINI}")
    log.info(f"Bulunan dosya: {len(xlsx_files)}")

    # Otomatik algılama
    HAFTALAR = hafta_algila(xlsx_files)
    TEMSILCILER = temsilci_algila(xlsx_files)

    log.info(f"Algılanan haftalar ({len(HAFTALAR)}):")
    for label, ws, we, _, _ in HAFTALAR:
        log.info(f"  - {label} ({ws} - {we})")

    log.info(f"Algılanan temsilciler ({len(TEMSILCILER)}):")
    for t in TEMSILCILER:
        log.info(f"  - {t}")

    DOSYA_TIPLERI = ['Planlanan Ziyaret', 'Yapılan Ziyaret', 'Sipariş Formu']

    # Veri toplama yapıları
    all_planlanan = []
    all_yapilan = []
    all_siparis = []
    customers = {}
    file_status = {}
    issues = []
    visits_seen = set()

    log.info("Dosyalar okunuyor...")
    for fn in xlsx_files:
        try:
            wbf = openpyxl.load_workbook(DOSYA_DIZINI / fn, data_only=True)
            wsf = wbf.active
            fn_norm = normalize_tr(fn)
            rep = temsilci_bul(fn, TEMSILCILER)
            week = hafta_bul(fn, HAFTALAR)

            if not rep or not week:
                log.warning(f"  ATLA: {fn} (temsilci={rep or '?'}, hafta={week or '?'})")
                wbf.close()
                continue

            ftype = dosya_tipi_bul(fn)
            is_planlanan = ftype == 'Planlanan Ziyaret'
            is_yapilan = ftype == 'Yapılan Ziyaret'
            is_siparis_file = ftype == 'Sipariş Formu'

            # Header satırı ve gömülü sipariş kontrolü
            has_embedded_siparis = False
            header_row = None

            for hr in range(1, 6):
                vals = ' '.join([safe_str(wsf.cell(row=hr, column=c).value) for c in range(1, wsf.max_column + 1)])
                vn = normalize_tr(vals)
                if 'URUN' in vn and ('FIYAT' in vn or 'ADET' in vn):
                    has_embedded_siparis = True
                    header_row = hr
                    break
                if ('LOKASYON' in vn or 'MUSTERI' in vn or 'GORUSULEN' in vn) and header_row is None:
                    header_row = hr

            if header_row is None:
                header_row = 3 if wsf.max_row > 3 else 1

            # Kolon eşleştirme
            cols = {}
            for c in range(1, wsf.max_column + 1):
                v = safe_str(wsf.cell(row=header_row, column=c).value)
                vn = normalize_tr(v)
                if ('MUSTERI' in vn or 'GORUSULEN' in vn or 'FIRMA' in vn) and 'musteri' not in cols:
                    cols['musteri'] = c
                elif ('YETKILI' in vn or 'KISI ADI' in vn) and 'yetkili' not in cols:
                    cols['yetkili'] = c
                elif ('ILETISIM' in vn or 'NUMARA' in vn or 'TELEFON' in vn) and 'iletisim' not in cols:
                    cols['iletisim'] = c
                elif ('LOKASYON' in vn or vn.strip() in ('IL', 'IL)')) and 'lokasyon' not in cols:
                    cols['lokasyon'] = c
                elif 'TARIH' in vn and 'GORUSME' not in vn and 'tarih' not in cols:
                    cols['tarih'] = c
                elif 'GUN' in vn and 'gun' not in cols:
                    cols['gun'] = c
                elif 'SURE' in vn and 'sure' not in cols:
                    cols['sure'] = c
                elif 'URUN' in vn and 'urun' not in cols:
                    cols['urun'] = c
                elif 'ADET' in vn and 'adet' not in cols:
                    cols['adet'] = c
                elif 'FIYAT' in vn and 'fiyat' not in cols:
                    cols['fiyat'] = c
                elif 'NOT' in vn and 'NUMARA' not in vn and 'notlar' not in cols:
                    cols['notlar'] = c

            # Dosya durumu kaydet
            if is_planlanan:
                file_status[(rep, week, 'Planlanan Ziyaret')] = fn
            if is_yapilan:
                file_status[(rep, week, 'Yapılan Ziyaret')] = fn
            if is_siparis_file:
                file_status[(rep, week, 'Sipariş Formu')] = fn
            if has_embedded_siparis and is_yapilan:
                file_status[(rep, week, 'Sipariş Formu')] = fn + ' (gömülü)'

            # Veri satırlarını oku
            if 'musteri' not in cols:
                log.warning(f"  UYARI: {fn} - 'Müşteri' kolonu bulunamadı, atlanıyor")
                wbf.close()
                continue

            last_musteri = last_yetkili = last_iletisim = last_tarih = ''

            for rr in range(header_row + 1, wsf.max_row + 1):
                # Tamamen boş satırları atla (trailing empty rows)
                row_has_data = any(
                    wsf.cell(row=rr, column=c).value
                    for c in range(1, min(wsf.max_column + 1, 12))
                )
                if not row_has_data:
                    continue

                musteri_val = safe_str(wsf.cell(row=rr, column=cols['musteri']).value)
                if musteri_val:
                    # Yeni müşteri bloğu: carry-forward değerlerini sıfırla
                    if musteri_val != last_musteri:
                        last_tarih = ''  # Tarih önceki müşteriden taşınmasın
                    last_musteri = musteri_val
                    last_yetkili = safe_str(wsf.cell(row=rr, column=cols.get('yetkili', 0)).value) if cols.get('yetkili') else ''
                    last_iletisim = safe_str(wsf.cell(row=rr, column=cols.get('iletisim', 0)).value) if cols.get('iletisim') else ''
                else:
                    musteri_val = last_musteri

                if not musteri_val:
                    continue

                yetkili = safe_str(wsf.cell(row=rr, column=cols.get('yetkili', 0)).value) if cols.get('yetkili') else ''
                if not yetkili:
                    yetkili = last_yetkili
                iletisim = safe_str(wsf.cell(row=rr, column=cols.get('iletisim', 0)).value) if cols.get('iletisim') else ''
                if not iletisim:
                    iletisim = last_iletisim

                tarih_val = wsf.cell(row=rr, column=cols.get('tarih', 0)).value if cols.get('tarih') else None
                tarih_str = safe_date(tarih_val)
                if tarih_str:
                    last_tarih = tarih_str
                else:
                    tarih_str = last_tarih

                lokasyon = safe_str(wsf.cell(row=rr, column=cols.get('lokasyon', 0)).value) if cols.get('lokasyon') else ''
                gun = safe_str(wsf.cell(row=rr, column=cols.get('gun', 0)).value) if cols.get('gun') else ''
                sure = safe_str(wsf.cell(row=rr, column=cols.get('sure', 0)).value) if cols.get('sure') else ''
                notlar = safe_str(wsf.cell(row=rr, column=cols.get('notlar', 0)).value) if cols.get('notlar') else ''
                urun = safe_str(wsf.cell(row=rr, column=cols.get('urun', 0)).value) if cols.get('urun') else ''
                adet = safe_str(wsf.cell(row=rr, column=cols.get('adet', 0)).value) if cols.get('adet') else ''
                fiyat = safe_str(wsf.cell(row=rr, column=cols.get('fiyat', 0)).value) if cols.get('fiyat') else ''

                # Müşteri master
                cust_key = normalize_tr(musteri_val)
                if cust_key and cust_key not in customers:
                    customers[cust_key] = {'name': musteri_val, 'yetkili': yetkili, 'iletisim': iletisim,
                                           'lokasyon': lokasyon, 'rep': rep, 'tarih': tarih_str}
                elif cust_key:
                    if yetkili and not customers[cust_key]['yetkili']:
                        customers[cust_key]['yetkili'] = yetkili
                    if iletisim and not customers[cust_key]['iletisim']:
                        customers[cust_key]['iletisim'] = iletisim
                    if tarih_str:
                        customers[cust_key]['tarih'] = tarih_str

                # Veri tipine göre kaydet
                if is_planlanan:
                    all_planlanan.append({'rep': rep, 'week': week, 'lokasyon': lokasyon, 'musteri': musteri_val,
                                          'tarih': tarih_str, 'gun': gun, 'notlar': notlar, 'dosya': fn})
                elif is_yapilan or is_siparis_file:
                    if is_yapilan:
                        # Müşteri+tarih+dosya bazında dedup (aynı müşterinin ürün alt-satırları tek ziyaret)
                        visit_key = (rep, normalize_tr(musteri_val), tarih_str, fn)
                        if visit_key not in visits_seen:
                            visits_seen.add(visit_key)
                            all_yapilan.append({'rep': rep, 'week': week, 'lokasyon': lokasyon, 'musteri': musteri_val,
                                                'yetkili': yetkili, 'iletisim': iletisim, 'tarih': tarih_str,
                                                'gun': gun, 'sure': sure, 'notlar': notlar, 'dosya': fn})

                    if urun or (is_siparis_file and (fiyat or adet)):
                        if not urun:
                            urun = '(Detaysız)'
                        all_siparis.append({'rep': rep, 'week': week, 'musteri': musteri_val, 'yetkili': yetkili,
                                            'iletisim': iletisim, 'tarih': tarih_str, 'urun': urun,
                                            'adet': adet, 'fiyat': fiyat, 'dosya': fn})

            wbf.close()
            log.info(f"  OK: {fn}")
        except Exception as e:
            log.error(f"  HATA: {fn} - {e}")

    log.info(f"Toplam: {len(all_planlanan)} planlanan, {len(all_yapilan)} yapılan, {len(all_siparis)} sipariş, {len(customers)} müşteri")

    # =============================================
    # EXCEL OLUŞTUR
    # =============================================
    wb = openpyxl.Workbook()

    # -- SHEET 1: DOSYA ENVANTER --
    ws1 = wb.active
    ws1.title = 'Dosya Envanteri'
    ws1.sheet_properties.tabColor = '1F4E79'
    headers1 = ['Temsilci', 'Hafta', 'Dosya Tipi', 'Dosya Adı', 'Durum', 'Kayıt Sayısı', 'Notlar']
    widths1 = [22, 20, 20, 55, 12, 14, 40]
    for i, (h, w) in enumerate(zip(headers1, widths1), 1):
        ws1.cell(row=1, column=i, value=h)
        ws1.column_dimensions[get_column_letter(i)].width = w
    style_header(ws1, 1, len(headers1))

    eksik_plan, eksik_yap, eksik_sip = [], [], []
    row = 2
    for rep in TEMSILCILER:
        for wlabel, _, _, _, _ in HAFTALAR:
            for ftype in DOSYA_TIPLERI:
                key = (rep, wlabel, ftype)
                found = file_status.get(key)
                if found:
                    if ftype == 'Planlanan Ziyaret':
                        cnt = len([r for r in all_planlanan if r['rep'] == rep and r['week'] == wlabel])
                    elif ftype == 'Yapılan Ziyaret':
                        cnt = len([r for r in all_yapilan if r['rep'] == rep and r['week'] == wlabel])
                    else:
                        cnt = len([r for r in all_siparis if r['rep'] == rep and r['week'] == wlabel])
                    status, fill, notes = 'MEVCUT', OK_FILL, f'{cnt} kayıt' if cnt else 'Veri yok'
                else:
                    cnt, status, fill = 0, 'EKSİK', ERR_FILL
                    notes = 'MANUEL GİRİŞ GEREKLİ'
                    if ftype == 'Planlanan Ziyaret':
                        eksik_plan.append((rep, wlabel))
                    elif ftype == 'Yapılan Ziyaret':
                        eksik_yap.append((rep, wlabel))
                    else:
                        eksik_sip.append((rep, wlabel))

                ws1.cell(row=row, column=1, value=rep)
                ws1.cell(row=row, column=2, value=wlabel)
                ws1.cell(row=row, column=3, value=ftype)
                ws1.cell(row=row, column=4, value=found or '-')
                ws1.cell(row=row, column=5, value=status)
                ws1.cell(row=row, column=6, value=cnt if cnt else '-')
                ws1.cell(row=row, column=7, value=notes)
                style_row(ws1, row, len(headers1), fill)
                row += 1
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(headers1))}{row - 1}'

    # -- SHEET 2: HAFTALIK TAKVİM --
    ws2 = wb.create_sheet('Haftalık Takvim')
    ws2.sheet_properties.tabColor = '2E7D32'
    headers2 = ['Hafta', 'Gün', 'Tarih'] + [f'{r}\n(Ziyaret)' for r in TEMSILCILER]
    widths2 = [20, 14, 14] + [30] * len(TEMSILCILER)
    for i, (h, w) in enumerate(zip(headers2, widths2), 1):
        ws2.cell(row=1, column=i, value=h)
        ws2.column_dimensions[get_column_letter(i)].width = w
    style_header(ws2, 1, len(headers2))

    visit_by_date = defaultdict(list)
    for v in all_planlanan + all_yapilan:
        if v['tarih']:
            visit_by_date[(v['tarih'], v['rep'])].append(v['musteri'])

    row2 = 2
    for wlabel, wstart, wend, sd, ed in HAFTALAR:
        num_days = (ed - sd).days + 1
        for day_offset in range(num_days):
            current_date = sd + timedelta(days=day_offset)
            if current_date.weekday() > 4:
                continue  # Hafta sonu atla
            date_str = current_date.strftime('%d.%m.%Y')
            gun = TR_DAYS.get(current_date.weekday(), '')

            ws2.cell(row=row2, column=1, value=wlabel)
            ws2.cell(row=row2, column=2, value=gun)
            ws2.cell(row=row2, column=3, value=date_str)

            for idx, rep in enumerate(TEMSILCILER):
                visits = visit_by_date.get((date_str, rep), [])
                seen = set()
                unique = []
                for m in visits:
                    mk = normalize_tr(m)
                    if mk not in seen:
                        seen.add(mk)
                        unique.append(m)
                cell_val = ', '.join(unique[:5])
                if len(unique) > 5:
                    cell_val += f' (+{len(unique) - 5})'
                ws2.cell(row=row2, column=4 + idx, value=cell_val if cell_val else '')

            fill = BLUE_FILL if day_offset % 2 == 0 else None
            style_row(ws2, row2, len(headers2), fill)
            row2 += 1
        # Haftalar arası boşluk
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=row2, column=c, value='')
        row2 += 1

    # -- SHEET 3: PLANLANAN ZİYARETLER --
    ws_p = wb.create_sheet('Planlanan Ziyaretler')
    ws_p.sheet_properties.tabColor = '3F51B5'
    p_headers = ['Temsilci', 'Hafta', 'Lokasyon', 'Müşteri', 'Ziyaret Tarihi', 'Gün', 'Notlar', 'Kaynak']
    p_widths = [20, 18, 16, 30, 14, 14, 40, 30]
    for i, (h, w) in enumerate(zip(p_headers, p_widths), 1):
        ws_p.cell(row=1, column=i, value=h)
        ws_p.column_dimensions[get_column_letter(i)].width = w
    style_header(ws_p, 1, len(p_headers))
    for rr_idx, rec in enumerate(all_planlanan, 2):
        for ci, key in enumerate(['rep', 'week', 'lokasyon', 'musteri', 'tarih', 'gun', 'notlar', 'dosya'], 1):
            ws_p.cell(row=rr_idx, column=ci, value=rec[key])
        style_row(ws_p, rr_idx, len(p_headers), AUTO_FILL if rr_idx % 2 == 0 else None)
    ws_p.auto_filter.ref = f'A1:{get_column_letter(len(p_headers))}{max(2, len(all_planlanan) + 1)}'

    # -- SHEET 4: YAPILAN ZİYARETLER --
    ws_y = wb.create_sheet('Yapılan Ziyaretler')
    ws_y.sheet_properties.tabColor = '2E7D32'
    y_headers = ['Temsilci', 'Hafta', 'Lokasyon', 'Müşteri', 'Yetkili Kişi', 'İletişim No', 'Ziyaret Tarihi', 'Gün', 'Görüşme Süresi', 'Notlar', 'Kaynak']
    y_widths = [20, 18, 16, 30, 22, 18, 14, 14, 12, 40, 30]
    for i, (h, w) in enumerate(zip(y_headers, y_widths), 1):
        ws_y.cell(row=1, column=i, value=h)
        ws_y.column_dimensions[get_column_letter(i)].width = w
    style_header(ws_y, 1, len(y_headers))
    y_keys = ['rep', 'week', 'lokasyon', 'musteri', 'yetkili', 'iletisim', 'tarih', 'gun', 'sure', 'notlar', 'dosya']
    for rr_idx, rec in enumerate(all_yapilan, 2):
        for ci, key in enumerate(y_keys, 1):
            ws_y.cell(row=rr_idx, column=ci, value=rec[key])
        style_row(ws_y, rr_idx, len(y_headers), PatternFill('solid', fgColor='F2FBF2') if rr_idx % 2 == 0 else None)
    ws_y.auto_filter.ref = f'A1:{get_column_letter(len(y_headers))}{max(2, len(all_yapilan) + 1)}'

    # -- SHEET 5: SİPARİŞLER --
    ws_s = wb.create_sheet('Siparişler')
    ws_s.sheet_properties.tabColor = 'C62828'
    s_headers = ['Temsilci', 'Hafta', 'Müşteri', 'Yetkili Kişi', 'İletişim No', 'Sipariş Tarihi', 'Ürün Adı', 'Adet', 'Fiyat', 'Kaynak']
    s_widths = [20, 18, 28, 22, 18, 14, 35, 12, 16, 30]
    for i, (h, w) in enumerate(zip(s_headers, s_widths), 1):
        ws_s.cell(row=1, column=i, value=h)
        ws_s.column_dimensions[get_column_letter(i)].width = w
    style_header(ws_s, 1, len(s_headers))
    s_keys = ['rep', 'week', 'musteri', 'yetkili', 'iletisim', 'tarih', 'urun', 'adet', 'fiyat', 'dosya']
    for rr_idx, rec in enumerate(all_siparis, 2):
        for ci, key in enumerate(s_keys, 1):
            ws_s.cell(row=rr_idx, column=ci, value=rec[key])
        style_row(ws_s, rr_idx, len(s_headers), LIGHT_FILL if rr_idx % 2 == 0 else None)
    ws_s.auto_filter.ref = f'A1:{get_column_letter(len(s_headers))}{max(2, len(all_siparis) + 1)}'

    # -- SHEET 6: MÜŞTERİ MASTER --
    ws3 = wb.create_sheet('Müşteri Master')
    ws3.sheet_properties.tabColor = 'FF6600'
    headers3 = ['Müşteri / Firma Adı', 'Yetkili Kişi', 'İletişim No', 'Lokasyon (İl)', 'Atanan Temsilci', 'Son Ziyaret Tarihi', 'Notlar']
    widths3 = [35, 25, 20, 16, 22, 16, 40]
    for i, (h, w) in enumerate(zip(headers3, widths3), 1):
        ws3.cell(row=1, column=i, value=h)
        ws3.column_dimensions[get_column_letter(i)].width = w
    style_header(ws3, 1, len(headers3))
    row3 = 2
    for key in sorted(customers.keys()):
        c = customers[key]
        ws3.cell(row=row3, column=1, value=c['name'])
        ws3.cell(row=row3, column=2, value=c['yetkili'])
        ws3.cell(row=row3, column=3, value=c['iletisim'])
        ws3.cell(row=row3, column=4, value=c['lokasyon'])
        ws3.cell(row=row3, column=5, value=c['rep'])
        ws3.cell(row=row3, column=6, value=c['tarih'])
        ws3.cell(row=row3, column=7, value='')
        style_row(ws3, row3, len(headers3), LIGHT_FILL if row3 % 2 == 0 else None)
        row3 += 1
    ws3.auto_filter.ref = f'A1:{get_column_letter(len(headers3))}{row3 - 1}'

    # -- SHEET 7: ÖZET --
    ws_oz = wb.create_sheet('Özet')
    ws_oz.sheet_properties.tabColor = '9C27B0'
    ws_oz.cell(row=1, column=1, value='SAHA SATIŞ HAFTALIK ÖZET RAPORU').font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws_oz.cell(row=2, column=1, value=f'Oluşturulma: {datetime.now().strftime("%d.%m.%Y %H:%M")}').font = Font(name='Arial', size=10, color='666666')
    ws_oz.cell(row=3, column=1, value=f'Kaynak: {len(xlsx_files)} dosya | {len(HAFTALAR)} hafta | {len(TEMSILCILER)} temsilci').font = Font(name='Arial', size=10, color='666666')

    oz_headers = ['Temsilci', 'Hafta', 'Planlanan', 'Yapılan', 'Gerçekleşme %', 'Sipariş Satırı', 'Benzersiz Müşteri']
    oz_widths = [22, 20, 12, 12, 14, 14, 16]
    for i, (h, w) in enumerate(zip(oz_headers, oz_widths), 1):
        ws_oz.cell(row=5, column=i, value=h)
        ws_oz.column_dimensions[get_column_letter(i)].width = w
    style_header(ws_oz, 5, len(oz_headers))

    row_oz = 6
    for rep in TEMSILCILER:
        for wlabel, _, _, _, _ in HAFTALAR:
            plan_cnt = len([r for r in all_planlanan if r['rep'] == rep and r['week'] == wlabel])
            yap_cnt = len([r for r in all_yapilan if r['rep'] == rep and r['week'] == wlabel])
            sip_cnt = len([r for r in all_siparis if r['rep'] == rep and r['week'] == wlabel])
            mus_set = set(
                [normalize_tr(r['musteri']) for r in all_yapilan if r['rep'] == rep and r['week'] == wlabel and r['musteri']] +
                [normalize_tr(r['musteri']) for r in all_siparis if r['rep'] == rep and r['week'] == wlabel and r['musteri']]
            )
            pct = f'{(yap_cnt / plan_cnt * 100):.0f}%' if plan_cnt > 0 else '-'

            ws_oz.cell(row=row_oz, column=1, value=rep)
            ws_oz.cell(row=row_oz, column=2, value=wlabel)
            ws_oz.cell(row=row_oz, column=3, value=plan_cnt or '-')
            ws_oz.cell(row=row_oz, column=4, value=yap_cnt or '-')
            ws_oz.cell(row=row_oz, column=5, value=pct)
            ws_oz.cell(row=row_oz, column=6, value=sip_cnt or '-')
            ws_oz.cell(row=row_oz, column=7, value=len(mus_set) or '-')
            style_row(ws_oz, row_oz, len(oz_headers))
            row_oz += 1

    ws_oz.cell(row=row_oz, column=1, value='TOPLAM').font = B_FONT
    ws_oz.cell(row=row_oz, column=3, value=len(all_planlanan)).font = B_FONT
    ws_oz.cell(row=row_oz, column=4, value=len(all_yapilan)).font = B_FONT
    t_pct = f'{(len(all_yapilan) / len(all_planlanan) * 100):.0f}%' if all_planlanan else '-'
    ws_oz.cell(row=row_oz, column=5, value=t_pct).font = B_FONT
    ws_oz.cell(row=row_oz, column=6, value=len(all_siparis)).font = B_FONT
    style_row(ws_oz, row_oz, len(oz_headers), PURPLE_FILL)

    # -- SHEET 8: VERİ KALİTE SORUNLARI --
    ws4 = wb.create_sheet('Veri Kalite Sorunları')
    ws4.sheet_properties.tabColor = 'F44336'
    headers4 = ['Öncelik', 'Kategori', 'Sorun Açıklaması', 'Etkilenen Dosya', 'Çözüm Önerisi', 'Durum']
    widths4 = [12, 20, 50, 45, 50, 12]
    for i, (h, w) in enumerate(zip(headers4, widths4), 1):
        ws4.cell(row=1, column=i, value=h)
        ws4.column_dimensions[get_column_letter(i)].width = w
    style_header(ws4, 1, len(headers4))

    # Otomatik sorun tespiti
    auto_issues = []
    for rep in TEMSILCILER:
        for wlabel, _, _, _, _ in HAFTALAR:
            for ftype in DOSYA_TIPLERI:
                if (rep, wlabel, ftype) not in file_status:
                    auto_issues.append(('KRİTİK', 'Eksik Dosya', f'{rep} - {wlabel} {ftype} dosyası eksik',
                                        '', 'Ekipten iste veya manuel gir', 'BEKLIYOR'))

    # Tarih uyumsuzlukları
    for rec in all_yapilan + all_planlanan:
        if rec['tarih'] and rec['week']:
            for wlabel, ws_str, we_str, sd, ed in HAFTALAR:
                if rec['week'] == wlabel:
                    try:
                        parts = rec['tarih'].split('.')
                        rec_dt = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                        if rec_dt < sd or rec_dt > ed:
                            auto_issues.append(('YÜKSEK', 'Tarih Hatası',
                                                f'{rec["rep"]}: {rec["tarih"]} tarihi {wlabel} aralığı dışında',
                                                rec.get('dosya', ''), 'Tarihi kontrol et', 'BEKLIYOR'))
                    except (ValueError, IndexError):
                        pass
                    break

    # Dedup issues
    seen_issues = set()
    unique_issues = []
    for issue in auto_issues:
        key = (issue[1], issue[2][:50])
        if key not in seen_issues:
            seen_issues.add(key)
            unique_issues.append(issue)

    row4 = 2
    for priority, cat, desc, file, fix, status in unique_issues:
        ws4.cell(row=row4, column=1, value=priority)
        ws4.cell(row=row4, column=2, value=cat)
        ws4.cell(row=row4, column=3, value=desc)
        ws4.cell(row=row4, column=4, value=file)
        ws4.cell(row=row4, column=5, value=fix)
        ws4.cell(row=row4, column=6, value=status)
        if priority == 'KRİTİK':
            fill = ERR_FILL
        elif priority == 'YÜKSEK':
            fill = WARN_FILL
        else:
            fill = INFO_FILL
        style_row(ws4, row4, len(headers4), fill)
        row4 += 1
    ws4.auto_filter.ref = f'A1:{get_column_letter(len(headers4))}{max(2, row4 - 1)}'

    # KAYDET (dosya kilidi korumalı)
    output = DOSYA_DIZINI / CIKTI_DOSYA
    try:
        wb.save(output)
    except PermissionError:
        log.error(f"MASTER_DATA.xlsx kaydedilemedi! Dosya başka bir programda (Excel?) açık olabilir.")
        log.error(f"Lütfen dosyayı kapatıp tekrar deneyin.")
        # Geçici dosyaya yaz
        temp_output = DOSYA_DIZINI / f'MASTER_DATA_TEMP_{datetime.now().strftime("%H%M%S")}.xlsx'
        wb.save(temp_output)
        log.info(f"Geçici dosyaya kaydedildi: {temp_output.name}")
        return False

    log.info(f'{"=" * 60}')
    log.info(f'MASTER_DATA.xlsx oluşturuldu!')
    log.info(f'{"=" * 60}')
    log.info(f'  Algılanan: {len(HAFTALAR)} hafta, {len(TEMSILCILER)} temsilci')
    log.info(f'  Sheet 1: Dosya Envanteri ({len(TEMSILCILER)} temsilci x {len(HAFTALAR)} hafta x 3 tip)')
    log.info(f'  Sheet 2: Haftalık Takvim (OTOMATİK DOLU)')
    log.info(f'  Sheet 3: Planlanan Ziyaretler ({len(all_planlanan)} kayıt)')
    log.info(f'  Sheet 4: Yapılan Ziyaretler ({len(all_yapilan)} kayıt)')
    log.info(f'  Sheet 5: Siparişler ({len(all_siparis)} kayıt)')
    log.info(f'  Sheet 6: Müşteri Master ({len(customers)} müşteri)')
    log.info(f'  Sheet 7: Özet (Performans tablosu)')
    log.info(f'  Sheet 8: Veri Kalite Sorunları ({len(unique_issues)} sorun)')
    log.info(f'Eksik: {len(eksik_plan)} planlanan, {len(eksik_yap)} yapılan, {len(eksik_sip)} sipariş')
    return True


# ============================================================
# İZLEME MODU (--izle)
# ============================================================

def izleme_modu():
    """Klasördeki değişiklikleri izle, yeni dosya gelince otomatik çalıştır."""
    import time
    DEBOUNCE_SANIYE = 3  # Değişiklik algılandıktan sonra bekleme süresi

    log.info("=" * 60)
    log.info("DOSYA İZLEME MODU AKTİF")
    log.info(f"Dizin: {DOSYA_DIZINI}")
    log.info("Yeni .xlsx dosyası eklendiğinde MASTER_DATA otomatik güncellenir.")
    log.info("Durdurmak için Ctrl+C")
    log.info("=" * 60)

    last_files = set(dosya_listesi_al(DOSYA_DIZINI))
    log.info(f"Mevcut dosya sayısı: {len(last_files)}")
    log.info("Bekleniyor...")

    try:
        while True:
            time.sleep(5)  # 5 saniyede bir kontrol
            current_files = set(dosya_listesi_al(DOSYA_DIZINI))

            yeni = current_files - last_files
            silinen = last_files - current_files

            if yeni or silinen:
                # Debounce: dosya kopyalanırken birden fazla değişiklik olabilir
                log.info(f"Değişiklik algılandı, {DEBOUNCE_SANIYE}s bekleniyor...")
                time.sleep(DEBOUNCE_SANIYE)
                # Debounce sonrası tekrar oku (dosya kopyası tamamlanmış olabilir)
                current_files = set(dosya_listesi_al(DOSYA_DIZINI))
                yeni = current_files - last_files
                silinen = last_files - current_files

                if yeni:
                    log.info(f"YENİ DOSYA:")
                    for f in yeni:
                        log.info(f"  + {f}")
                if silinen:
                    log.info(f"SİLİNEN DOSYA:")
                    for f in silinen:
                        log.info(f"  - {f}")

                log.info("MASTER_DATA güncelleniyor...")
                master_data_olustur()
                last_files = current_files
                log.info("Tamamlandı. Bekleniyor...")

    except KeyboardInterrupt:
        log.info("İzleme durduruldu.")


if __name__ == '__main__':
    if '--izle' in sys.argv or '--watch' in sys.argv:
        # Önce bir kez çalıştır, sonra izle
        master_data_olustur()
        izleme_modu()
    else:
        master_data_olustur()
